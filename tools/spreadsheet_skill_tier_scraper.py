#!/usr/bin/env python3
"""Maintainer tool: ingest the community "Uma Musume Skills Spreadsheet" into
data/skill_tiers_normalized.json for the skill-purchase scorer (redesign Stage 3).

Source: the public Google Sheet's `alldata` tab (the flat master table). Each skill
carries a Team-Trials rank, a PvP/Champions rank, Score/SP, and base cost. Ranks are
the sheet's symbols (essential -> useless):  ⍟ > ◎ > ◯ > ▲ > △ > ✕.

Output JSON keyed by the SAME normalization the scorer uses (norm(strip_marks(name)))
so career_bot/skills.py can look skills up by name. Placeholder family rows
("[Run Style] Corners ◎", "[Distance] Straightaways ◯") are expanded into their
concrete style/distance variants. Names are validated against the game's skill list
(data/skill_condition_core.json); unmatched names are logged, not dropped.

Deps: openpyxl (and internet, unless --xlsx points at a local workbook).
Usage:
  python tools/spreadsheet_skill_tier_scraper.py            # download + write data/skill_tiers_normalized.json
  python tools/spreadsheet_skill_tier_scraper.py --xlsx C:\\path\\to\\sheet.xlsx
"""
import argparse
import json
import re
import sys
import urllib.request
from pathlib import Path

SHEET_ID = "1oB3eTvKqREtJDWJL0q80O_VjBcpOmRl5xE0z5fZKgFY"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
RANK_SYMBOLS = {"⍟", "◎", "◯", "▲", "△", "✕"}
_NAME_MARKS = "◎◯⍟▲△✕○●☆"
STYLE_WORDS = ["Front Runner", "Pace Chaser", "Late Surger", "End Closer"]
DISTANCE_WORDS = ["Sprint", "Mile", "Medium", "Long"]

BASE = Path(__file__).resolve().parent.parent


def strip_marks(s):
    s = str(s or "")
    for m in _NAME_MARKS:
        s = s.replace(m, "")
    return s.strip()


def norm_key(s):
    """Must mirror career_bot.skills.norm(strip_mark(...))."""
    return re.sub(r"[^a-z0-9]+", "", strip_marks(s).lower())


def expand_placeholder(name):
    """A '[Run Style] X' / '[Distance] X' family row -> concrete variant names."""
    out = []
    if "[Run Style]" in name:
        for w in STYLE_WORDS:
            out.append(name.replace("[Run Style]", w))
    if "[Distance]" in name:
        for w in DISTANCE_WORDS:
            out.append(name.replace("[Distance]", w))
    return out


def clean_rank(v):
    v = str(v or "").strip()
    return v if v in RANK_SYMBOLS else ""


def _load_alldata_rows(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "alldata" not in wb.sheetnames:
        raise SystemExit(f"'alldata' tab not found; tabs = {wb.sheetnames}")
    ws = wb["alldata"]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def build(xlsx_path, out_path):
    rows = _load_alldata_rows(xlsx_path)
    # alldata block 1 (the canonical schema): col1 name, col2 TT, col3 PvP,
    # col4 Score/SP, col5 Base Cost. (Blocks 2/3 duplicate it for display.)
    game_names = set()
    cond = json.loads((BASE / "data" / "skill_condition_core.json").read_text(encoding="utf-8"))
    for e in cond if isinstance(cond, list) else []:
        if e.get("name"):
            game_names.add(norm_key(e["name"]))

    tiers = {}
    unmatched = []
    placeholders = 0

    def add(name, tt, cm, score, cost):
        key = norm_key(name)
        if not key:
            return
        tt = clean_rank(tt)
        cm = clean_rank(cm)
        if not tt and not cm:
            return
        rec = tiers.get(key)
        # On a normalize-collision (e.g. ◎ vs ◯ variant) keep the stronger TT rank.
        order = {"⍟": 6, "◎": 5, "◯": 4, "▲": 3, "△": 2, "✕": 1, "": 0}
        if rec and order.get(tt, 0) <= order.get(rec.get("tt", ""), 0):
            return
        try:
            score_f = round(float(score), 4) if str(score or "").strip() else None
        except Exception:
            score_f = None
        tiers[key] = {
            "name": strip_marks(name),
            "tt": tt,
            "cm": cm,
            "score_per_sp": score_f,
            "base_cost": str(cost or "").strip(),
            "matched": key in game_names,
        }

    for r in rows[2:]:
        if len(r) < 6:
            continue
        name = str(r[1] or "").strip()
        if not name or name.lower() == "skill name":
            continue
        tt, cm, score, cost = r[2], r[3], r[4], r[5]
        if "[" in name:  # placeholder family row -> expand
            placeholders += 1
            for ex in expand_placeholder(name):
                add(ex, tt, cm, score, cost)
        else:
            add(name, tt, cm, score, cost)

    matched = sum(1 for v in tiers.values() if v["matched"])
    for v in tiers.values():
        if not v["matched"]:
            unmatched.append(v["name"])

    payload = {
        "_meta": {
            "source": "community Uma Musume Skills Spreadsheet (alldata tab)",
            "sheet_id": SHEET_ID,
            "rank_order": "essential>best>good>situational>excess>useless = " + " ".join("⍟◎◯▲△✕"),
            "ranks": {"tt": "Rank (Team Trials)", "cm": "Rank (PvP / Champions Meeting)"},
            "count": len(tiers),
            "matched_game_skills": matched,
            "placeholder_rows_expanded": placeholders,
        },
        "tiers": tiers,
    }
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"wrote {out_path}  ({len(tiers)} skills, {matched} matched game names, {placeholders} placeholder rows)")
    if unmatched:
        print(f"  {len(unmatched)} unmatched (kept anyway): {unmatched[:12]}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="", help="local .xlsx (default: download the sheet)")
    ap.add_argument("--out", default=str(BASE / "data" / "skill_tiers_normalized.json"))
    args = ap.parse_args()
    xlsx = args.xlsx
    tmp = None
    if not xlsx:
        tmp = BASE / "data" / "_sheet_tmp.xlsx"
        print("downloading sheet ...")
        urllib.request.urlretrieve(EXPORT_URL, tmp)
        xlsx = str(tmp)
    try:
        build(xlsx, Path(args.out))
    finally:
        if tmp and tmp.exists():
            tmp.unlink()


if __name__ == "__main__":
    main()
